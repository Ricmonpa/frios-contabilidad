# -*- coding: utf-8 -*-
"""
N3Labs | Agente Contable Frios - Conciliador Automático
====================================================
Demo de conciliación automática SAT (Dsoft) vs SAP Business One
y generación del Anexo 11 - IVA Acreditable (DIOT).

Ejecutar con:  streamlit run app.py
"""

import os
import time
import base64
import random
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st

# ---------------------------------------------------------------------------
# IDENTIDAD DE MARCA (Frios + N3Labs)
# ---------------------------------------------------------------------------
FRIOS_PURPLE = "#5D1F77"       # morado corporativo Frios (tomado del logo)
FRIOS_PURPLE_DARK = "#43155A"
FRIOS_PURPLE_LIGHT = "#8B4FA8"
BASE_DIR = Path(__file__).parent


def _img_b64(nombre: str) -> str:
    """Devuelve un data-URI base64 de una imagen del proyecto (o "" si falta)."""
    ruta = BASE_DIR / nombre
    if not ruta.exists():
        return ""
    data = base64.b64encode(ruta.read_bytes()).decode("utf-8")
    return f"data:image/png;base64,{data}"


LOGO_FRIOS = _img_b64("logo-frios.png")
LOGO_N3 = _img_b64("n3-logo.png")

# ---------------------------------------------------------------------------
# CONFIGURACIÓN GENERAL
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="N3Labs | Agente Contable Frios",
    page_icon="🧊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Estilo corporativo ligero
st.markdown(
    f"""
    <style>
    .main-title {{
        font-size: 2.0rem; font-weight: 800; color: {FRIOS_PURPLE};
        margin-bottom: 0;
    }}
    .subtitle {{ color: {FRIOS_PURPLE_LIGHT}; font-size: 1.05rem; margin-top: 0; }}
    div[data-testid="stMetric"] {{
        background: #FAF6FC; border: 1px solid #E4D4EC;
        border-left: 4px solid {FRIOS_PURPLE};
        border-radius: 12px; padding: 14px 18px;
    }}
    .stTabs [data-baseweb="tab"] {{ font-size: 1.02rem; font-weight: 600; }}
    .stTabs [aria-selected="true"] {{ color: {FRIOS_PURPLE} !important; }}
    div.stButton > button[kind="primary"] {{
        background: linear-gradient(90deg, {FRIOS_PURPLE_DARK} 0%, {FRIOS_PURPLE} 55%, {FRIOS_PURPLE_LIGHT} 100%);
        color: white; border: none; border-radius: 10px;
        padding: 0.65rem 1.4rem; font-weight: 700;
    }}
    div.stButton > button[kind="primary"]:hover {{
        filter: brightness(1.1); color: white;
    }}
    /* Encabezado con logo Frios */
    .brand-header {{
        display: flex; align-items: center; gap: 18px; margin-bottom: 2px;
    }}
    .brand-header img {{ height: 54px; border-radius: 8px; }}
    /* Bloque N3Labs en la barra lateral */
    .n3-block {{
        background: linear-gradient(135deg, {FRIOS_PURPLE_DARK} 0%, {FRIOS_PURPLE} 100%);
        border-radius: 12px; padding: 14px 16px; text-align: center;
        margin-top: 6px;
    }}
    .n3-block img {{ height: 52px; margin-top: 8px; }}
    .n3-block .n3-caption {{
        color: #EAD9F2; font-size: 0.82rem; letter-spacing: 0.3px; margin: 0;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------------
# DATOS DUMMY POR DEFECTO (formato basado en el export real de Dsoft:
# hoja "1.2 Cfdi Recibido Dsoft" del Anexo 11)
# ---------------------------------------------------------------------------
PROVEEDORES = [
    ("FFR860630K48", "FRIALSA FRIGORIFICOS SA DE CV"),
    ("FAC910401756", "BBVA LEASING MEXICO SA DE CV"),
    ("QUI840127F69", "QUIMPER SA DE CV"),
    ("TBU210421JV6", "TECH BUKERS SAPI DE CV"),
    ("ISD950921HE5", "PASE SERVICIOS ELECTRONICOS SA DE CV"),
    ("TUI220104SC0", "TUIO SAPI DE CV"),
    ("CSS160330CP7", "CFE SUMINISTRADOR DE SERVICIOS BASICOS"),
    ("TME840315KT6", "TELEFONOS DE MEXICO SAB DE CV"),
    ("GNP9211244P0", "GRUPO NACIONAL PROVINCIAL SAB"),
    ("FLE9309069V7", "FACILEASING SA DE CV"),
    ("SGA061013C82", "SERVICIOS GASOLINEROS DE MEXICO SA"),
    ("MTR120927E36", "MANTENIMIENTO Y REFRIGERACION DEL BAJIO SA"),
]


def _uuid_falso(rng: random.Random) -> str:
    """Genera un UUID con formato de folio fiscal del SAT."""
    hexc = "0123456789ABCDEF"
    partes = [8, 4, 4, 4, 12]
    return "-".join("".join(rng.choice(hexc) for _ in range(n)) for n in partes)


def generar_datos_sat() -> pd.DataFrame:
    """DataFrame dummy con el formato del export Dsoft/SAT (Anexo 11, hoja 1.2)."""
    rng = random.Random(2025)
    filas = []
    for i, (rfc, nombre) in enumerate(PROVEEDORES):
        base = round(rng.uniform(2500, 95000), 2)
        iva = round(base * 0.16, 2)
        filas.append(
            {
                "UUID": _uuid_falso(rng),
                "Serie": rng.choice(["FE", "A", "B", "CFDI"]),
                "Folio": str(rng.randint(100000, 999999)),
                "Fecha Emisión": f"2025-05-{rng.randint(2, 28):02d}",
                "RFC Emisor": rfc,
                "Proveedor": nombre,
                "Base IVA 16%": base,
                "IVA": iva,
                "Total": round(base + iva, 2),
                "Moneda": "MXN",
                "Método Pago": rng.choice(["PUE", "PPD"]),
            }
        )
    return pd.DataFrame(filas)


def normalizar_sat(df: pd.DataFrame) -> pd.DataFrame:
    """Mapea columnas de un export real de Dsoft al esquema interno del demo."""
    mapa = {
        "uuid": "UUID",
        "folio fiscal": "UUID",
        "serie": "Serie",
        "folio": "Folio",
        "folioexterno": "Folio",
        "emisión": "Fecha Emisión",
        "emision": "Fecha Emisión",
        "fecha": "Fecha Emisión",
        "fechacompra": "Fecha Emisión",
        "emisor rfc": "RFC Emisor",
        "rfc": "RFC Emisor",
        "emisor nombre": "Proveedor",
        "nombre": "Proveedor",
        "proveedor": "Proveedor",
        "base iva 16": "Base IVA 16%",
        "base iva 16%": "Base IVA 16%",
        "subtotal": "Base IVA 16%",
        "iva": "IVA",
        "total": "Total",
        "total factura": "Total",
        "moneda": "Moneda",
        "método pago descripción": "Método Pago",
        "metodo pago": "Método Pago",
        "metodopago": "Método Pago",
    }
    renombres = {}
    for col in df.columns:
        clave = str(col).strip().lower()
        if clave in mapa and mapa[clave] not in renombres.values():
            renombres[col] = mapa[clave]
    df = df.rename(columns=renombres)

    if "UUID" not in df.columns:
        raise ValueError("No se encontró la columna UUID en el archivo.")

    df = df[df["UUID"].notna() & (df["UUID"].astype(str).str.len() > 10)].copy()
    df = df.drop_duplicates(subset="UUID", keep="first")

    # Fechas en serial de Excel (p.ej. 46146 en exports .xlsb) → fecha legible
    if "Fecha Emisión" in df.columns:
        fechas = pd.to_numeric(df["Fecha Emisión"], errors="coerce")
        seriales = fechas.between(35000, 60000)
        if seriales.any():
            convertidas = pd.to_datetime(
                fechas[seriales], unit="D", origin="1899-12-30"
            ).dt.strftime("%Y-%m-%d")
            df["Fecha Emisión"] = df["Fecha Emisión"].astype(str)
            df.loc[seriales, "Fecha Emisión"] = convertidas
        df["Fecha Emisión"] = df["Fecha Emisión"].astype(str).str.slice(0, 10)

    for col in ["Base IVA 16%", "IVA", "Total"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0).round(2)
        else:
            df[col] = 0.0
    for col in ["Serie", "Folio", "Fecha Emisión", "RFC Emisor", "Proveedor",
                "Moneda", "Método Pago"]:
        if col not in df.columns:
            df[col] = ""
    orden = ["UUID", "Serie", "Folio", "Fecha Emisión", "RFC Emisor", "Proveedor",
             "Base IVA 16%", "IVA", "Total", "Moneda", "Método Pago"]
    return df[orden].reset_index(drop=True)


def leer_archivo_subido(archivo):
    """Lee CSV / XLSX / XLSB. Devuelve (df_normalizado, hoja_cruda, complementos).

    Si es el Anexo 11 real (.xlsb), busca la hoja con más facturas (la
    "1.3 Pagos con UUID"), conserva sus columnas originales (cuenta bancaria,
    pólizas, folios) y además lee la hoja de complementos de pago (1.4)."""
    nombre = archivo.name.lower()
    if nombre.endswith(".csv"):
        return normalizar_sat(pd.read_csv(archivo)), None, None

    engine = "pyxlsb" if nombre.endswith(".xlsb") else None
    xls = pd.ExcelFile(archivo, engine=engine)

    def _fila_encabezado(hoja, texto):
        try:
            crudo = xls.parse(hoja, header=None, nrows=15)
        except Exception:
            return None
        for i in range(len(crudo)):
            if any(str(v).strip().upper() == texto for v in crudo.iloc[i].tolist()):
                return i
        return None

    # Hoja principal: la que tenga más filas con UUID válido
    mejor, mejor_crudo, max_filas = None, None, 0
    for hoja in xls.sheet_names:
        fila_hdr = _fila_encabezado(hoja, "UUID")
        if fila_hdr is None:
            continue
        try:
            crudo = xls.parse(hoja, header=fila_hdr)
            datos = normalizar_sat(crudo)
        except Exception:
            continue
        if len(datos) > max_filas:
            mejor, mejor_crudo, max_filas = datos, crudo, len(datos)

    if mejor is None:
        raise ValueError("Ninguna hoja del archivo contiene una columna UUID.")

    # Hoja de complementos de pago (1.4): se identifica por la columna UUIDRel
    complementos = None
    for hoja in xls.sheet_names:
        if "complemento" not in hoja.lower():
            continue
        fila_hdr = _fila_encabezado(hoja, "UUIDREL")
        if fila_hdr is not None:
            try:
                complementos = xls.parse(hoja, header=fila_hdr)
            except Exception:
                complementos = None
        break

    return mejor, mejor_crudo, complementos


def generar_mock_sap(df_sat: pd.DataFrame) -> pd.DataFrame:
    """Simula la extracción de SAP Business One a partir de los datos del SAT.
    Idéntico ~90%, pero introduce discrepancias intencionales para el demo:
      - 2 facturas que existen en SAT pero NO están registradas en SAP
      - 2 facturas con diferencia de centavos en el IVA capturado
    """
    rng = random.Random(42)
    df = df_sat.copy()
    n = len(df)

    # Solo sembrar discrepancias en facturas con IVA real (> $10)
    candidatos = [i for i in range(n) if df_sat.loc[i, "IVA"] > 10] or list(range(n))
    rng.shuffle(candidatos)
    faltantes = candidatos[:2] if len(candidatos) >= 6 else candidatos[:1]
    con_diferencia = candidatos[2:4] if len(candidatos) >= 6 else candidatos[1:2]

    df = df.drop(index=faltantes).reset_index(drop=True)

    uuids_diferencia = df_sat.loc[con_diferencia, "UUID"].tolist()
    for u in uuids_diferencia:
        mask = df["UUID"] == u
        ajuste = round(rng.choice([-1, 1]) * rng.uniform(0.15, 9.90), 2)
        df.loc[mask, "IVA"] = (df.loc[mask, "IVA"] + ajuste).round(2)
        df.loc[mask, "Total"] = (df.loc[mask, "Base IVA 16%"] + df.loc[mask, "IVA"]).round(2)

    # Darle "sabor" SAP B1 al extracto (como la hoja 1.3 Pagos con UUID)
    df_sap = pd.DataFrame(
        {
            "DocEntry": [2400000 + rng.randint(1000, 99999) for _ in range(len(df))],
            "UUID": df["UUID"],
            "CardCode": ["P" + str(rng.randint(10, 999)).zfill(4) for _ in range(len(df))],
            "RFC": df["RFC Emisor"],
            "Proveedor": df["Proveedor"],
            "Fecha Contab.": df["Fecha Emisión"],
            "Subtotal": df["Base IVA 16%"],
            "IVA": df["IVA"],
            "Total Factura": df["Total"],
            "Cuenta": "11801000",
            "Nombre Cuenta": "IVA Acreditable",
            "Estatus": "ACTIVA",
        }
    )
    return df_sap


# ---------------------------------------------------------------------------
# PAPEL DE TRABAJO — ANEXO 11 (formato Frios, hoja "IVA")
# Réplica de la estructura real: 1.1 estados de cuenta · 1.2 CFDI · 1.3 SAP ·
# 1.4 complementos de pago, con cruce y diferencia por factura.
# ---------------------------------------------------------------------------
COLS_PAPEL_BANCO = ["BANCO", "# MOVIMIENTO", "REF", "CONCEPTO", "REFERENCIA AMPLIADA",
                    "IMPORTE", "FECHA PAGO BANCO", "CRUCE"]
COLS_PAPEL_CFDI = ["RFC", "NOMBRE", "UUID FOLIO FISCAL", "REGIMEN FISCAL",
                   "FECHA COMPROBANTE", "CONCEPTO DEL BIEN O SERVICIO",
                   "BASE 16%", "BASE 8%", "BASE 0%", "BASE EXENTA", "IVA", "IEPS",
                   "RET IVA", "RET ISR", "TOTAL CFDI", "MONEDA",
                   "FORMA DE PAGO", "METODO PAGO"]
COLS_PAPEL_SAP = ["NÚMERO DE PROVEEDOR", "FOLIO EXTERNO", "NÚMERO DE PAGO SISTEMA",
                  "FECHA DE PAGO SISTEMA", "TOTAL PAGO", "TIPO",
                  "NÚMERO PÓLIZA SISTEMA", "FECHA PÓLIZA", "CUENTA CONTABLE",
                  "NOMBRE CUENTA", "DIF", "COMENTARIOS"]
COLS_PAPEL_CP = ["COMPLEMENTO DE PAGO (UUID)", "MONTO DEL PAGO", "FECHA COMPLEMENTO",
                 "BASE 16% (CP)", "IVA (CP)", "TOTAL (CP)"]
COLS_PAPEL = COLS_PAPEL_BANCO + COLS_PAPEL_CFDI + COLS_PAPEL_SAP + COLS_PAPEL_CP
COLS_PAPEL_DINERO = ["IMPORTE", "BASE 16%", "BASE 8%", "BASE 0%", "BASE EXENTA",
                     "IVA", "IEPS", "RET IVA", "RET ISR", "TOTAL CFDI", "TOTAL PAGO",
                     "DIF", "MONTO DEL PAGO", "BASE 16% (CP)", "IVA (CP)", "TOTAL (CP)"]


def construir_papel_trabajo(merge: pd.DataFrame, crudo: pd.DataFrame | None,
                            complementos: pd.DataFrame | None) -> pd.DataFrame:
    """Arma el papel de trabajo del Anexo 11 fila por fila.

    Con archivo real usa las columnas originales de la hoja 1.3 (cuenta,
    pólizas, folios) y los complementos reales de la 1.4; en modo demo
    sintetiza esos datos de forma verosímil."""
    rng = random.Random(7)
    bancos = ["Bancomer", "Banorte", "Santander"]

    # Índices por UUID para cruzar contra las hojas reales
    idx_crudo, cols_crudo = None, {}
    if crudo is not None and "UUID" in [str(c).strip().upper() for c in crudo.columns]:
        cols_crudo = {str(c).strip().upper(): c for c in crudo.columns}
        idx_crudo = crudo.copy()
        idx_crudo["__uuid"] = idx_crudo[cols_crudo["UUID"]].astype(str).str.strip().str.upper()
        idx_crudo = idx_crudo.drop_duplicates("__uuid").set_index("__uuid")

    idx_cp, cols_cp = None, {}
    if complementos is not None:
        cols_cp = {str(c).strip().upper(): c for c in complementos.columns}
        if "UUIDREL" in cols_cp:
            idx_cp = complementos.copy()
            idx_cp["__uuid"] = idx_cp[cols_cp["UUIDREL"]].astype(str).str.strip().str.upper()
            idx_cp = idx_cp.drop_duplicates("__uuid").set_index("__uuid")

    def _crudo_val(uuid, *nombres):
        if idx_crudo is None or uuid not in idx_crudo.index:
            return None
        for n in nombres:
            col = cols_crudo.get(n.upper())
            if col is not None:
                v = idx_crudo.at[uuid, col]
                if pd.notna(v) and str(v).strip() != "":
                    return v
        return None

    def _num(v):
        v = pd.to_numeric(v, errors="coerce") if v is not None else None
        return round(float(v), 2) if v is not None and pd.notna(v) else None

    filas = []
    solo_sat = merge[merge["_merge"] != "right_only"].reset_index(drop=True)
    for i, f in solo_sat.iterrows():
        uuid = str(f["UUID"]).strip().upper()
        falta_en_sap = f["_merge"] == "left_only"
        con_diferencia = f["Estatus Conciliación"] == "⚠️ Diferencia de monto"
        es_ppd = "PPD" in str(f["Método Pago"]).upper() or "PARCIALIDADES" in str(f["Método Pago"]).upper()

        banco = str(_crudo_val(uuid, "NOMBRE CUENTA") or rng.choice(bancos))
        fecha_pago = _crudo_val(uuid, "FECHA") or f["Fecha Emisión"]
        total_pago = _num(_crudo_val(uuid, "TOTAL PAGO")) or f["Total"]
        folio_ext = _crudo_val(uuid, "FOLIOEXTERNO") or f"{f['Serie']}{f['Folio']}"

        fila = {
            # ---- 1.1 Estados de cuenta ----
            "BANCO": banco,
            "# MOVIMIENTO": str(rng.randint(200000, 990000)) if crudo is None else "",
            "REF": str(rng.randint(100000, 999999)) if crudo is None else "",
            "CONCEPTO": str(f["Proveedor"])[:28],
            "REFERENCIA AMPLIADA": f"PAGO FACTURA {folio_ext}",
            "IMPORTE": -round(float(total_pago), 2) if not falta_en_sap else None,
            "FECHA PAGO BANCO": str(fecha_pago)[:10] if not falta_en_sap else "",
            "CRUCE": f"{banco[:3].upper()}{i + 1}" if not falta_en_sap else "",
            # ---- 1.2 CFDI recibidos (SAT / Dsoft) ----
            "RFC": f["RFC Emisor"],
            "NOMBRE": f["Proveedor"],
            "UUID FOLIO FISCAL": f["UUID"],
            "REGIMEN FISCAL": str(_crudo_val(uuid, "REGIMEN") or "601"),
            "FECHA COMPROBANTE": str(f["Fecha Emisión"])[:10],
            "CONCEPTO DEL BIEN O SERVICIO": f"BIENES Y SERVICIOS {str(f['Proveedor'])[:20]}",
            "BASE 16%": f["Base IVA 16%"],
            "BASE 8%": 0.0, "BASE 0%": 0.0, "BASE EXENTA": 0.0,
            "IVA": f["IVA"], "IEPS": 0.0,
            "RET IVA": _num(_crudo_val(uuid, "RETENCION")) or 0.0,
            "RET ISR": 0.0,
            "TOTAL CFDI": f["Total"],
            "MONEDA": str(f["Moneda"]).replace("$", "MXN") or "MXN",
            "FORMA DE PAGO": "Transferencia electrónica",
            "METODO PAGO": "PPD" if es_ppd else "PUE",
        }

        # ---- 1.3 SAP Business One ----
        if falta_en_sap:
            fila.update({c: (None if c in ("TOTAL PAGO", "DIF") else "") for c in COLS_PAPEL_SAP})
            fila["COMENTARIOS"] = "NO REGISTRADA EN SAP — registrar antes del cierre"
        else:
            dif = round(float(f["Dif. IVA"]), 2) if pd.notna(f["Dif. IVA"]) else 0.0
            fila.update({
                "NÚMERO DE PROVEEDOR": str(_crudo_val(uuid, "CLIENTE") or f"P{rng.randint(10, 999):04d}"),
                "FOLIO EXTERNO": str(folio_ext),
                "NÚMERO DE PAGO SISTEMA": str(_crudo_val(uuid, "PAGO") or rng.randint(48000, 48999)),
                "FECHA DE PAGO SISTEMA": str(fecha_pago)[:10],
                "TOTAL PAGO": round(float(total_pago), 2),
                "TIPO": str(_crudo_val(uuid, "TIPO") or "FACTURA"),
                "NÚMERO PÓLIZA SISTEMA": str(_crudo_val(uuid, "ASIENTO TT")
                                             or f.get("DocEntry") or rng.randint(2400000, 2499999)),
                "FECHA PÓLIZA": str(_crudo_val(uuid, "FECHA TT") or fecha_pago)[:10],
                "CUENTA CONTABLE": str(_crudo_val(uuid, "CUENTA") or "11102002"),
                "NOMBRE CUENTA": banco,
                "DIF": dif,
                "COMENTARIOS": ("Corregir captura de IVA en SAP" if con_diferencia else "OK"),
            })

        # ---- 1.4 Complemento de pago (solo facturas PPD) ----
        cp_uuid, cp_monto, cp_fecha, cp_base, cp_iva, cp_total = "", None, "", None, None, None
        if es_ppd and not falta_en_sap:
            if idx_cp is not None:
                if uuid in idx_cp.index:
                    g = idx_cp.loc[uuid]
                    col = lambda n: cols_cp.get(n.upper())
                    cp_uuid = str(g[col("UUID")]) if col("UUID") else ""
                    cp_monto = _num(g[col("MONTO")]) if col("MONTO") else None
                    cp_fecha = str(g[col("FECHA")])[:10] if col("FECHA") else ""
                    cp_base = _num(g[col("BASE IVA 16% TRASLADO")]) if col("BASE IVA 16% TRASLADO") else None
                    cp_iva = _num(g[col("IVA 16% TRASLADO")]) if col("IVA 16% TRASLADO") else None
                    cp_total = _num(g[col("IMPPAGADO")]) if col("IMPPAGADO") else cp_monto
            else:  # modo demo: sintetizar el complemento
                cp_uuid = _uuid_falso(rng)
                cp_monto = cp_total = round(float(total_pago), 2)
                cp_fecha = str(fecha_pago)[:10]
                cp_base = f["Base IVA 16%"]
                cp_iva = f["IVA"]
        fila.update({
            "COMPLEMENTO DE PAGO (UUID)": cp_uuid,
            "MONTO DEL PAGO": cp_monto,
            "FECHA COMPLEMENTO": cp_fecha,
            "BASE 16% (CP)": cp_base,
            "IVA (CP)": cp_iva,
            "TOTAL (CP)": cp_total,
        })
        filas.append(fila)

    return pd.DataFrame(filas, columns=COLS_PAPEL)


def exportar_excel_papel(resultado: dict) -> bytes:
    """Genera el Excel del papel de trabajo con el layout del Anexo 11 real:
    títulos, bloques 1.1–1.4, encabezados y fila de TOTALES (DIOT)."""
    import io
    from openpyxl.styles import Font, PatternFill, Alignment

    papel = resultado["papel"]
    morado = "5D1F77"
    lila = "EADAF2"
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        papel.to_excel(w, sheet_name="IVA (Papel de Trabajo)", startrow=4, index=False)
        ws = w.sheets["IVA (Papel de Trabajo)"]

        ws["A1"] = "FRIOS"
        ws["A1"].font = Font(bold=True, size=14, color=morado)
        ws["A2"] = "Papel de Trabajo — Anexo 11 · IVA Acreditable (DIOT)"
        ws["A2"].font = Font(bold=True, size=11)
        ws["A3"] = "Generado automáticamente por el Agente Contable · N3Labs"
        ws["A3"].font = Font(italic=True, size=9, color="777777")

        # Fila de bloques (1.1 – 1.4) con celdas combinadas
        bloques = [
            ("1.1  ESTADOS DE CUENTA", len(COLS_PAPEL_BANCO)),
            ("1.2  CFDI RECIBIDOS (SAT / DSOFT)", len(COLS_PAPEL_CFDI)),
            ("1.3  SAP BUSINESS ONE", len(COLS_PAPEL_SAP)),
            ("1.4  COMPLEMENTOS DE PAGO", len(COLS_PAPEL_CP)),
        ]
        col = 1
        for titulo, ancho in bloques:
            c = ws.cell(row=4, column=col, value=titulo)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=morado)
            c.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + ancho - 1)
            col += ancho

        # Encabezados
        for j in range(1, len(COLS_PAPEL) + 1):
            c = ws.cell(row=5, column=j)
            c.font = Font(bold=True, size=9, color=morado)
            c.fill = PatternFill("solid", fgColor=lila)

        # Fila TOTALES (como el renglón "ACREDITAMIENTO DIOT" del papel real)
        fila_tot = 5 + len(papel) + 1
        ws.cell(row=fila_tot, column=1, value="TOTALES — ACREDITAMIENTO DIOT").font = Font(bold=True, color=morado)
        for nombre in ["BASE 16%", "IVA", "RET IVA", "RET ISR", "TOTAL CFDI", "TOTAL PAGO"]:
            j = papel.columns.get_loc(nombre) + 1
            total = pd.to_numeric(papel[nombre], errors="coerce").sum()
            c = ws.cell(row=fila_tot, column=j, value=round(float(total), 2))
            c.font = Font(bold=True, color=morado)
            c.fill = PatternFill("solid", fgColor=lila)
            c.number_format = "#,##0.00"

        for j, nombre in enumerate(COLS_PAPEL, start=1):
            letra = ws.cell(row=5, column=j).column_letter
            ws.column_dimensions[letra].width = 34 if "UUID" in nombre else 15
            if nombre in COLS_PAPEL_DINERO:
                for r in range(6, 6 + len(papel)):
                    ws.cell(row=r, column=j).number_format = "#,##0.00"

        # Hojas complementarias
        cols_disc = ["Estatus Conciliación", "UUID", "Proveedor", "RFC Emisor",
                     "IVA", "IVA SAP", "Dif. IVA", "Total", "Total SAP"]
        resultado["discrepancias"][cols_disc].to_excel(w, sheet_name="Discrepancias", index=False)
        resultado["anexo11"].to_excel(w, sheet_name="DIOT (Anexo 11)", index=False)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# MÓDULO 2 (VISTA PREVIA) — CONCILIACIÓN BANCARIA DE EGRESOS
# Simula la ingesta del estado de cuenta (formato hoja "1.1 BBVA" del papel
# real) y el cruce automático banco ↔ SAP ↔ CFDI con CRUZE y DIF.
# ---------------------------------------------------------------------------
def generar_estado_cuenta_mock(papel: pd.DataFrame) -> pd.DataFrame:
    """Estado de cuenta simulado a partir de los pagos del papel de trabajo,
    con discrepancias intencionales: comisiones sin CFDI y un pago de SAP
    que no aparece en el banco."""
    rng = random.Random(11)
    pagados = papel[papel["IMPORTE"].notna()].reset_index(drop=True)

    filas = []
    for i, f in pagados.iterrows():
        filas.append({
            "Banco": f["BANCO"],
            "No. Movimiento": str(f["# MOVIMIENTO"]) or str(rng.randint(200000, 990000)),
            "Fecha Operación": f["FECHA PAGO BANCO"],
            "Concepto": f["CONCEPTO"],
            "Referencia Ampliada": f["REFERENCIA AMPLIADA"],
            "Cargo": f["IMPORTE"],
            "UUID": f["UUID FOLIO FISCAL"],
        })

    # Un pago registrado en SAP que NO aparece en el banco
    if len(filas) > 4:
        filas.pop(rng.randrange(len(filas)))

    # Comisiones bancarias sin factura (cargos sin CFDI)
    for concepto, monto in [("COMISION MEMBRESIA", -348.00),
                            ("COMISION TRANSFERENCIA SPEI", -87.00),
                            ("IVA COMISIONES", -69.60)]:
        filas.append({
            "Banco": rng.choice(["Bancomer", "Banorte"]),
            "No. Movimiento": str(rng.randint(200000, 990000)),
            "Fecha Operación": "2025-05-30",
            "Concepto": concepto,
            "Referencia Ampliada": concepto,
            "Cargo": monto,
            "UUID": "",
        })
    return pd.DataFrame(filas)


def cruzar_banco(estado: pd.DataFrame, papel: pd.DataFrame) -> dict:
    """Cruce automático banco ↔ SAP ↔ CFDI (la columna CRUZE/DIF del papel
    real, que hoy se hace a mano)."""
    pagados = papel[papel["IMPORTE"].notna()].set_index("UUID FOLIO FISCAL")
    uuids_banco = set(estado["UUID"][estado["UUID"].astype(str).str.len() > 10])

    filas = []
    for _, mov in estado.iterrows():
        uuid = str(mov["UUID"])
        if len(uuid) > 10 and uuid in pagados.index:
            p = pagados.loc[uuid]
            dif = round(abs(float(mov["Cargo"])) - float(p["TOTAL PAGO"]), 2)
            filas.append({
                "Banco": mov["Banco"], "Fecha": mov["Fecha Operación"],
                "Concepto": mov["Concepto"], "Cargo": mov["Cargo"],
                "# Pago SAP": p["NÚMERO DE PAGO SISTEMA"],
                "Póliza": p["NÚMERO PÓLIZA SISTEMA"],
                "Proveedor": p["NOMBRE"],
                "CRUZE": abs(mov["Cargo"]),
                "DIF": dif,
                "Estatus": "✅ Cruzado" if abs(dif) < 0.01 else "⚠️ Diferencia",
            })
        else:
            filas.append({
                "Banco": mov["Banco"], "Fecha": mov["Fecha Operación"],
                "Concepto": mov["Concepto"], "Cargo": mov["Cargo"],
                "# Pago SAP": "", "Póliza": "", "Proveedor": "",
                "CRUZE": None, "DIF": None,
                "Estatus": "❌ Cargo sin CFDI",
            })

    # Pagos de SAP que no aparecen en el banco
    sin_banco = pagados[~pagados.index.isin(uuids_banco)].reset_index()
    for _, p in sin_banco.iterrows():
        filas.append({
            "Banco": p["BANCO"], "Fecha": p["FECHA DE PAGO SISTEMA"],
            "Concepto": p["CONCEPTO"], "Cargo": None,
            "# Pago SAP": p["NÚMERO DE PAGO SISTEMA"],
            "Póliza": p["NÚMERO PÓLIZA SISTEMA"],
            "Proveedor": p["NOMBRE"],
            "CRUZE": None, "DIF": None,
            "Estatus": "🔎 Pago SAP sin reflejo en banco",
        })

    tabla = pd.DataFrame(filas)
    return {
        "tabla": tabla,
        "n_cruzados": int((tabla["Estatus"] == "✅ Cruzado").sum()),
        "n_sin_cfdi": int((tabla["Estatus"] == "❌ Cargo sin CFDI").sum()),
        "n_sin_banco": int((tabla["Estatus"] == "🔎 Pago SAP sin reflejo en banco").sum()),
        "n_diferencias": int((tabla["Estatus"] == "⚠️ Diferencia").sum()),
    }


# ---------------------------------------------------------------------------
# LÓGICA DE CONCILIACIÓN
# ---------------------------------------------------------------------------
def ejecutar_conciliacion(df_sat: pd.DataFrame, df_sap: pd.DataFrame,
                          crudo: pd.DataFrame | None = None,
                          complementos: pd.DataFrame | None = None) -> dict:
    """Cruza SAT vs SAP por UUID y clasifica cada factura."""
    sap = df_sap[["UUID", "DocEntry", "Subtotal", "IVA", "Total Factura"]].rename(
        columns={"Subtotal": "Base SAP", "IVA": "IVA SAP", "Total Factura": "Total SAP"}
    )
    m = df_sat.merge(sap, on="UUID", how="outer", indicator=True)

    def clasificar(fila):
        if fila["_merge"] == "left_only":
            return "❌ Falta en SAP"
        if fila["_merge"] == "right_only":
            return "⚠️ Solo en SAP"
        if abs(fila["IVA"] - fila["IVA SAP"]) > 0.005 or abs(fila["Total"] - fila["Total SAP"]) > 0.005:
            return "⚠️ Diferencia de monto"
        return "✅ Conciliada"

    m["Estatus Conciliación"] = m.apply(clasificar, axis=1)
    m["Dif. IVA"] = (m["IVA"].fillna(0) - m["IVA SAP"].fillna(0)).round(2)

    discrepancias = m[m["Estatus Conciliación"] != "✅ Conciliada"].copy()
    discrepancias["Acción sugerida"] = discrepancias["Estatus Conciliación"].map(
        {
            "❌ Falta en SAP": "Registrar factura en SAP B1 antes del cierre",
            "⚠️ Diferencia de monto": "Corregir captura de IVA en SAP (dif. de centavos)",
            "⚠️ Solo en SAP": "Validar CFDI contra el SAT / posible cancelación",
        }
    )

    conciliadas = m[m["Estatus Conciliación"] == "✅ Conciliada"]

    # ---- Anexo 11 - IVA Acreditable (DIOT), consolidado por proveedor ----
    base = m[m["_merge"] != "right_only"].copy()
    anexo = (
        base.groupby(["RFC Emisor", "Proveedor"], as_index=False)
        .agg(
            **{
                "Núm. Facturas": ("UUID", "count"),
                "Base Gravada 16%": ("Base IVA 16%", "sum"),
                "IVA Acreditable 16%": ("IVA", "sum"),
                "Total Operaciones": ("Total", "sum"),
            }
        )
        .round(2)
    )
    anexo.insert(0, "Tipo de Tercero", "04 - Proveedor Nacional")
    anexo.insert(1, "Tipo de Operación", "85 - Otros")
    pendientes = set(discrepancias["RFC Emisor"].dropna())
    anexo["Estatus DIOT"] = anexo["RFC Emisor"].map(
        lambda r: "⚠️ Revisar" if r in pendientes else "✅ Listo"
    )
    anexo = anexo.sort_values("IVA Acreditable 16%", ascending=False).reset_index(drop=True)

    return {
        "merge": m,
        "conciliadas": conciliadas,
        "discrepancias": discrepancias,
        "anexo11": anexo,
        "papel": construir_papel_trabajo(m, crudo, complementos),
        "n_conciliadas": len(conciliadas),
        "n_faltantes": int((m["Estatus Conciliación"] == "❌ Falta en SAP").sum()),
        "n_diferencias": int((m["Estatus Conciliación"] == "⚠️ Diferencia de monto").sum()),
        "iva_acreditable": float(conciliadas["IVA"].sum()),
        "iva_en_riesgo": float(discrepancias["IVA"].fillna(0).sum()),
    }


# ---------------------------------------------------------------------------
# AGENTE DE CHAT
# Por defecto responde con reglas sobre la conciliación (sin costo).
# Si existe GEMINI_API_KEY en st.secrets, usa la API de Gemini con el
# contexto de la conciliación; si algo falla, cae a las reglas.
# ---------------------------------------------------------------------------
def _secreto(nombre: str, default: str = "") -> str:
    """Lee un secreto de st.secrets (Streamlit Cloud / local) o de variables
    de entorno (Railway, Render, etc.)."""
    try:
        valor = st.secrets.get(nombre, "")
        if valor:
            return valor
    except Exception:  # sin archivo secrets.toml
        pass
    return os.environ.get(nombre, default)


def _gemini_api_key() -> str:
    return _secreto("GEMINI_API_KEY")


def _contexto_conciliacion() -> str:
    """Resumen en texto de la conciliación para dárselo como contexto al LLM."""
    r = st.session_state.get("resultado")
    if r is None:
        return "Aún no se ha ejecutado la conciliación de este mes."
    disc = r["discrepancias"]
    lineas = [
        f"- {f['Estatus Conciliación']} | {f['Proveedor']} (RFC {f['RFC Emisor']}) | "
        f"IVA SAT ${f['IVA'] if pd.notna(f['IVA']) else 0:,.2f} | "
        f"IVA SAP ${f['IVA SAP'] if pd.notna(f['IVA SAP']) else 0:,.2f} | UUID {f['UUID']}"
        for _, f in disc.iterrows()
    ]
    return (
        f"Facturas totales: {len(r['merge'])}. Conciliadas: {r['n_conciliadas']}. "
        f"Faltantes en SAP: {r['n_faltantes']}. Diferencias de monto: {r['n_diferencias']}. "
        f"IVA acreditable conciliado: ${r['iva_acreditable']:,.2f}. "
        f"IVA en revisión: ${r['iva_en_riesgo']:,.2f}.\n"
        "Discrepancias detectadas:\n" + "\n".join(lineas)
    )


def responder_con_gemini(pregunta: str) -> str | None:
    """Intenta responder con Gemini. Devuelve None si no hay clave o falla."""
    api_key = _gemini_api_key()
    if not api_key:
        return None
    try:
        from google import genai

        modelo = _secreto("GEMINI_MODEL", "gemini-2.5-flash")
        client = genai.Client(api_key=api_key)
        prompt = (
            "Eres el Agente Contable de Frios, un asistente de N3Labs que ayuda al "
            "equipo de finanzas (Lucila y Sergio) a revisar la conciliación mensual "
            "entre las facturas del SAT (descargadas con Dsoft) y SAP Business One, "
            "para el Anexo 11 - IVA Acreditable (DIOT). Responde en español, breve y "
            "profesional, en Markdown, escapando los signos de pesos como \\$. "
            "Básate únicamente en estos datos de la conciliación:\n\n"
            f"{_contexto_conciliacion()}\n\nPregunta del usuario: {pregunta}"
        )
        respuesta = client.models.generate_content(model=modelo, contents=prompt)
        return respuesta.text
    except Exception:
        return None  # sin conexión, sin paquete o error de API → usar reglas


def responder_agente(pregunta: str) -> str:
    respuesta_llm = responder_con_gemini(pregunta)
    if respuesta_llm:
        return respuesta_llm
    return responder_con_reglas(pregunta)


def responder_con_reglas(pregunta: str) -> str:
    q = pregunta.lower()
    r = st.session_state.get("resultado")

    if r is None:
        return (
            "Aún no he ejecutado la conciliación de este mes. 🙌\n\n"
            "Ve a la pestaña **2. Conciliación Inteligente** y presiona "
            "**Ejecutar Conciliación con IA**; en cuanto termine podré responder "
            "sobre facturas faltantes, diferencias de montos y el Anexo 11."
        )

    disc = r["discrepancias"]
    faltantes = disc[disc["Estatus Conciliación"] == "❌ Falta en SAP"]
    diferencias = disc[disc["Estatus Conciliación"] == "⚠️ Diferencia de monto"]

    # --- ¿Qué facturas faltan en SAP? ---
    if any(k in q for k in ["falta", "faltan", "faltantes", "no están", "no estan", "missing"]):
        if faltantes.empty:
            return "¡Buenas noticias! ✅ No detecté facturas del SAT que falten en SAP este mes."
        lineas = [
            f"- **{f['Proveedor']}** (RFC {f['RFC Emisor']}) por **\\${f['Total']:,.2f}** "
            f"con IVA de \\${f['IVA']:,.2f} — UUID `{f['UUID']}`"
            for _, f in faltantes.iterrows()
        ]
        return (
            f"Detecté **{len(faltantes)} facturas** que están timbradas en el SAT (Dsoft) "
            "pero **no encontré su registro en SAP Business One**:\n\n"
            + "\n".join(lineas)
            + "\n\n💡 *Recomendación: registrarlas en SAP antes del cierre para no perder "
            f"**\\${faltantes['IVA'].sum():,.2f}** de IVA acreditable.*"
        )

    # --- Diferencias de montos ---
    if any(k in q for k in ["diferencia", "monto", "centavo", "descuadre", "no cuadra"]):
        if diferencias.empty:
            return "No hay diferencias de montos este mes: SAT y SAP cuadran al centavo. ✅"
        lineas = [
            f"- **{f['Proveedor']}**: IVA en SAT \\${f['IVA']:,.2f} vs SAP \\${f['IVA SAP']:,.2f} "
            f"(diferencia de **\\${abs(f['Dif. IVA']):,.2f}**) — UUID `{f['UUID']}`"
            for _, f in diferencias.iterrows()
        ]
        return (
            f"Encontré **{len(diferencias)} facturas con diferencias de monto** entre "
            "Dsoft/SAT y SAP:\n\n" + "\n".join(lineas)
            + "\n\n💡 *Suelen ser errores de captura o redondeo. Sugiero corregir el "
            "asiento en SAP para que el Anexo 11 cuadre exacto.*"
        )

    # --- Totales / IVA acreditable ---
    if any(k in q for k in ["total", "iva acreditable", "cuánto", "cuanto", "suma", "acreditar"]):
        return (
            "Resumen del mes 📊\n\n"
            f"- **IVA acreditable conciliado:** \\${r['iva_acreditable']:,.2f}\n"
            f"- **IVA en revisión (discrepancias):** \\${r['iva_en_riesgo']:,.2f}\n"
            f"- **Facturas conciliadas:** {r['n_conciliadas']} de {len(r['merge'])}\n\n"
            "Si resolvemos las discrepancias, el IVA acreditable total sería de "
            f"**\\${r['iva_acreditable'] + r['iva_en_riesgo']:,.2f}**."
        )

    # --- Papel de trabajo ---
    if any(k in q for k in ["papel", "columnas", "banco", "complemento", "devoluci"]):
        papel = r.get("papel")
        n_cp = int((papel["COMPLEMENTO DE PAGO (UUID)"].astype(str).str.len() > 10).sum()) if papel is not None else 0
        return (
            "El **Papel de Trabajo del Anexo 11** ya está generado en la pestaña 2, "
            "con la estructura completa del formato de Frios:\n\n"
            "- **1.1 Estados de cuenta** (banco, movimiento, importe, cruce)\n"
            "- **1.2 CFDI recibidos** (UUID, bases por tasa, IVA, retenciones)\n"
            "- **1.3 SAP Business One** (pago, póliza, cuenta contable, DIF)\n"
            "- **1.4 Complementos de pago** (UUID del CP, monto, base e IVA)\n\n"
            f"Incluye {len(papel) if papel is not None else 0} facturas y {n_cp} complementos de pago vinculados. "
            "Puedes descargarlo en **Excel** con la fila de TOTALES (acreditamiento DIOT), "
            "listo para la relación de comprobación de devoluciones de IVA. 📄"
        )

    # --- Anexo 11 / DIOT ---
    if any(k in q for k in ["anexo", "diot", "reporte", "declaración", "declaracion"]):
        top = r["anexo11"].head(3)
        lineas = [
            f"- {t['Proveedor']}: IVA \\${t['IVA Acreditable 16%']:,.2f} ({t['Núm. Facturas']} facturas)"
            for _, t in top.iterrows()
        ]
        return (
            "El **Anexo 11 - IVA Acreditable (DIOT)** ya está pre-generado en la pestaña 2. "
            f"Consolida **{len(r['anexo11'])} proveedores**. Los de mayor IVA acreditable:\n\n"
            + "\n".join(lineas)
            + "\n\n⚠️ Los proveedores marcados con *Revisar* tienen discrepancias pendientes."
        )

    # --- Proveedores ---
    if "proveedor" in q:
        pend = disc["Proveedor"].dropna().unique().tolist()
        return (
            "Los proveedores con pendientes este mes son: "
            + ", ".join(f"**{p}**" for p in pend)
            + ". El resto está conciliado al 100%. ✅"
        )

    # --- Default ---
    return (
        f"Puedo ayudarte a analizar la conciliación de este mes "
        f"({r['n_faltantes'] + r['n_diferencias']} discrepancias detectadas). "
        "Prueba preguntarme:\n\n"
        "- *¿Qué facturas faltan en SAP?*\n"
        "- *¿Cuáles tienen diferencias de montos?*\n"
        "- *¿Cuánto IVA puedo acreditar este mes?*\n"
        "- *¿Cómo va el Anexo 11 / DIOT?*"
    )


# ---------------------------------------------------------------------------
# ESTADO INICIAL
# ---------------------------------------------------------------------------
if "df_sat" not in st.session_state:
    st.session_state.df_sat = generar_datos_sat()
    st.session_state.origen_sat = "Datos de demostración (formato Dsoft/SAT)"
if "df_sap" not in st.session_state:
    st.session_state.df_sap = None
if "crudo_sat" not in st.session_state:
    st.session_state.crudo_sat = None
if "cp_sat" not in st.session_state:
    st.session_state.cp_sat = None
if "resultado" not in st.session_state:
    st.session_state.resultado = None
if "chat" not in st.session_state:
    st.session_state.chat = [
        {
            "role": "assistant",
            "content": (
                "Hola Lucila/Sergio. He analizado la conciliación de este mes. "
                "Encontré 3 diferencias entre Dsoft y SAP. ¿Qué te gustaría revisar?"
            ),
        }
    ]

# ---------------------------------------------------------------------------
# SIDEBAR
# ---------------------------------------------------------------------------
with st.sidebar:
    if LOGO_FRIOS:
        st.markdown(
            f'<img src="{LOGO_FRIOS}" alt="Frios" style="height:46px;border-radius:8px;">',
            unsafe_allow_html=True,
        )
    st.markdown("**Agente Contable**")
    st.caption(f"Cliente: **Frios** · Periodo: **Mayo 2025** · {datetime.now():%d/%m/%Y}")
    st.divider()
    st.markdown(
        "**Flujo del demo:**\n"
        "1. Sube el export de Dsoft (o usa los datos demo)\n"
        "2. Conecta a SAP Business One\n"
        "3. Ejecuta la conciliación con IA\n"
        "4. Pregúntale al Agente Contable 💬"
    )
    st.divider()
    _n3_img = f'<img src="{LOGO_N3}" alt="N3Labs">' if LOGO_N3 else ""
    st.markdown(
        f'<div class="n3-block">'
        f'<p class="n3-caption">Hecho por</p>'
        f'{_n3_img}</div>',
        unsafe_allow_html=True,
    )
    st.caption("© 2025 N3Labs · Demo confidencial")

# ---------------------------------------------------------------------------
# ENCABEZADO
# ---------------------------------------------------------------------------
_logo_img = f'<img src="{LOGO_FRIOS}" alt="Frios">' if LOGO_FRIOS else ""
st.markdown(
    f'<div class="brand-header">{_logo_img}'
    f'<p class="main-title">Agente Contable - Conciliador Automático ✨</p></div>',
    unsafe_allow_html=True,
)
st.markdown('<p class="subtitle">Conciliación automática SAT (Dsoft) vs SAP Business One · Anexo 11 - IVA Acreditable (DIOT)</p>', unsafe_allow_html=True)
st.divider()

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📥 1. Integración de Datos",
    "🤖 2. Conciliación Inteligente",
    "💬 3. Chat Contable (Agente)",
    "🏦 4. Conciliación Bancaria · Módulo 2",
    "📈 5. Ingresos SAP↔Banca · Módulo 3",
])

# ===========================================================================
# TAB 1 — INTEGRACIÓN DE DATOS
# ===========================================================================
with tab1:
    col_izq, col_der = st.columns(2, gap="large")

    with col_izq:
        st.subheader("📄 Facturas SAT (export Dsoft)")
        st.caption("Sube el export de CFDIs recibidos (CSV, XLSX o el Anexo 11 en XLSB).")
        archivo = st.file_uploader(
            "Upload Dsoft / SAT XML Export",
            type=["csv", "xlsx", "xls", "xlsb"],
            help="Acepta el archivo 'Anexo 11 - IVA Acreditable' tal como lo genera Dsoft.",
        )
        if archivo is not None:
            try:
                df_leido, crudo_leido, cp_leido = leer_archivo_subido(archivo)
                st.session_state.df_sat = df_leido
                st.session_state.crudo_sat = crudo_leido
                st.session_state.cp_sat = cp_leido
                st.session_state.origen_sat = f"Archivo cargado: {archivo.name}"
                st.session_state.df_sap = None
                st.session_state.resultado = None
                st.session_state.banco = None
                st.success(f"✅ {len(st.session_state.df_sat)} CFDIs cargados desde **{archivo.name}**")
            except Exception as e:
                st.error(f"No pude interpretar el archivo ({e}). Usando datos de demostración.")

        st.info(f"**Fuente actual:** {st.session_state.origen_sat}", icon="🗂️")
        st.dataframe(
            st.session_state.df_sat,
            width="stretch",
            height=320,
            column_config={
                "Base IVA 16%": st.column_config.NumberColumn(format="$%.2f"),
                "IVA": st.column_config.NumberColumn(format="$%.2f"),
                "Total": st.column_config.NumberColumn(format="$%.2f"),
            },
        )
        st.metric("CFDIs del SAT cargados", len(st.session_state.df_sat))

    with col_der:
        st.subheader("🏢 SAP Business One")
        st.caption("Extracción de movimientos de proveedores vía Service Layer API.")

        if st.button("🔌 Conectar a SAP Business One (Vía API)", type="primary", width="stretch"):
            with st.spinner("Conectando a SAP B1... Extrayendo movimientos..."):
                time.sleep(2.5)
                st.session_state.df_sap = generar_mock_sap(st.session_state.df_sat)
                st.session_state.resultado = None
            st.toast("Conexión exitosa con SAP Business One", icon="✅")

        if st.session_state.df_sap is not None:
            st.success(
                f"✅ Conectado a **SAP B1 - FRIOS_PROD** · "
                f"{len(st.session_state.df_sap)} movimientos extraídos (cuenta 11801000 - IVA Acreditable)"
            )
            st.dataframe(
                st.session_state.df_sap,
                width="stretch",
                height=320,
                column_config={
                    "Subtotal": st.column_config.NumberColumn(format="$%.2f"),
                    "IVA": st.column_config.NumberColumn(format="$%.2f"),
                    "Total Factura": st.column_config.NumberColumn(format="$%.2f"),
                },
            )
            st.metric("Movimientos SAP extraídos", len(st.session_state.df_sap))
        else:
            st.warning("Aún no hay conexión con SAP. Haz clic en el botón para extraer los movimientos.", icon="🔌")

# ===========================================================================
# TAB 2 — CONCILIACIÓN INTELIGENTE Y ANEXO 11
# ===========================================================================
with tab2:
    st.subheader("🤖 Conciliación Inteligente SAT vs SAP")

    if st.session_state.df_sap is None:
        st.warning("Primero conecta a SAP Business One en la pestaña **1. Integración de Datos**.", icon="⚠️")
    else:
        if st.button("✨ EJECUTAR CONCILIACIÓN CON IA ✨", type="primary", width="stretch"):
            with st.spinner("El agente está cruzando UUIDs, validando montos e identificando discrepancias..."):
                time.sleep(2.0)
                st.session_state.resultado = ejecutar_conciliacion(
                    st.session_state.df_sat, st.session_state.df_sap,
                    st.session_state.crudo_sat, st.session_state.cp_sat,
                )
                st.session_state.banco = None
            st.toast("Conciliación completada", icon="🎉")

    r = st.session_state.resultado
    if r is not None:
        st.divider()
        c1, c2, c3 = st.columns(3)
        c1.metric(
            "✅ Facturas Conciliadas",
            r["n_conciliadas"],
            delta=f"${r['iva_acreditable']:,.2f} IVA acreditable",
        )
        c2.metric(
            "❌ Faltantes en SAP",
            r["n_faltantes"],
            delta="Requieren registro",
            delta_color="inverse",
        )
        c3.metric(
            "⚠️ Diferencias de Montos",
            r["n_diferencias"],
            delta="Errores de captura",
            delta_color="inverse",
        )

        st.divider()
        st.markdown("### 🚨 Discrepancias que requieren atención")
        st.caption("El agente detectó estas facturas al cruzar Dsoft/SAT contra SAP B1 por UUID.")
        cols_disc = [
            "Estatus Conciliación", "UUID", "Proveedor", "RFC Emisor",
            "IVA", "IVA SAP", "Dif. IVA", "Total", "Total SAP", "Acción sugerida",
        ]
        st.dataframe(
            r["discrepancias"][cols_disc].rename(columns={"IVA": "IVA SAT", "Total": "Total SAT"}),
            width="stretch",
            column_config={
                "IVA SAT": st.column_config.NumberColumn(format="$%.2f"),
                "IVA SAP": st.column_config.NumberColumn(format="$%.2f"),
                "Dif. IVA": st.column_config.NumberColumn(format="$%.2f"),
                "Total SAT": st.column_config.NumberColumn(format="$%.2f"),
                "Total SAP": st.column_config.NumberColumn(format="$%.2f"),
            },
        )

        st.divider()
        st.markdown("### 📋 Pre-visualización Anexo 11 - IVA Acreditable (DIOT)")
        st.caption("Formato consolidado por proveedor, listo para revisión del contador.")
        st.dataframe(
            r["anexo11"],
            width="stretch",
            column_config={
                "Base Gravada 16%": st.column_config.NumberColumn(format="$%.2f"),
                "IVA Acreditable 16%": st.column_config.NumberColumn(format="$%.2f"),
                "Total Operaciones": st.column_config.NumberColumn(format="$%.2f"),
            },
        )

        tot1, tot2, tot3 = st.columns(3)
        tot1.metric("Base Gravada 16% Total", f"${r['anexo11']['Base Gravada 16%'].sum():,.2f}")
        tot2.metric("IVA Acreditable Total", f"${r['anexo11']['IVA Acreditable 16%'].sum():,.2f}")
        tot3.metric("Proveedores en DIOT", len(r["anexo11"]))

        csv = r["anexo11"].to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "⬇️ Descargar Anexo 11 (CSV)",
            data=csv,
            file_name="Anexo11_IVA_Acreditable_DIOT.csv",
            mime="text/csv",
        )

        st.divider()
        st.markdown("### 📄 Papel de Trabajo — Anexo 11 (formato Frios)")
        st.caption(
            "Estructura del papel de trabajo real, factura por factura: "
            "**1.1 estados de cuenta · 1.2 CFDI recibidos · 1.3 SAP Business One · "
            "1.4 complementos de pago**, con cruce y diferencia. Es la base de la DIOT "
            "y de la relación de comprobación para devoluciones de IVA."
        )
        st.dataframe(
            r["papel"],
            width="stretch",
            height=420,
            column_config={
                c: st.column_config.NumberColumn(format="$%.2f")
                for c in COLS_PAPEL_DINERO
            },
        )
        st.download_button(
            "⬇️ Descargar Papel de Trabajo completo (Excel)",
            data=exportar_excel_papel(r),
            file_name="Anexo11_Papel_de_Trabajo_IVA_Acreditable.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

# ===========================================================================
# TAB 3 — CHAT CONTABLE (AGENTE FASE 2)
# ===========================================================================
with tab3:
    st.subheader("💬 Chat Contable — Agente IA")
    st.caption(
        "Pregunta en lenguaje natural sobre la conciliación: facturas faltantes, "
        "diferencias de montos, IVA acreditable o el Anexo 11."
    )

    for msg in st.session_state.chat:
        avatar = "🧊" if msg["role"] == "assistant" else "👤"
        with st.chat_message(msg["role"], avatar=avatar):
            st.markdown(msg["content"])

    pregunta = st.chat_input("Escribe tu pregunta... ej. ¿Qué facturas faltan en SAP?")
    if pregunta:
        st.session_state.chat.append({"role": "user", "content": pregunta})
        with st.chat_message("user", avatar="👤"):
            st.markdown(pregunta)
        with st.chat_message("assistant", avatar="🧊"):
            with st.spinner("Analizando la conciliación..."):
                time.sleep(1.2)
                respuesta = responder_agente(pregunta)
            st.markdown(respuesta)
        st.session_state.chat.append({"role": "assistant", "content": respuesta})

# ===========================================================================
# TAB 4 — CONCILIACIÓN BANCARIA (MÓDULO 2 · VISTA PREVIA)
# ===========================================================================
with tab4:
    st.subheader("🏦 Conciliación Bancaria de Egresos")
    st.markdown(
        f'<span style="background:{FRIOS_PURPLE};color:white;padding:3px 12px;'
        f'border-radius:12px;font-size:0.8rem;font-weight:700;">VISTA PREVIA · MÓDULO 2 DE LA RUTA INTEGRAL</span>',
        unsafe_allow_html=True,
    )
    st.caption(
        "Ingesta de estados de cuenta multi-banco y cruce automático "
        "**banco ↔ SAP ↔ CFDI** — las columnas CRUZE y DIF del papel de trabajo, "
        "que hoy se llenan a mano, generadas por el agente."
    )

    if st.session_state.resultado is None:
        st.warning("Primero ejecuta la conciliación SAT vs SAP en la pestaña **2. Conciliación Inteligente**.", icon="⚠️")
    else:
        if st.button("🏦 Cargar Estados de Cuenta y Cruzar (simulado)", type="primary", width="stretch"):
            with st.spinner("Leyendo estados de cuenta... Cruzando movimientos contra SAP y CFDIs..."):
                time.sleep(2.0)
                estado = generar_estado_cuenta_mock(st.session_state.resultado["papel"])
                st.session_state.banco = cruzar_banco(estado, st.session_state.resultado["papel"])
            st.toast("Cruce bancario completado", icon="🏦")

        b = st.session_state.get("banco")
        if b is not None:
            st.divider()
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("✅ Movimientos Cruzados", b["n_cruzados"])
            m2.metric("❌ Cargos sin CFDI", b["n_sin_cfdi"], delta="Comisiones / sin factura", delta_color="inverse")
            m3.metric("🔎 Pagos SAP sin banco", b["n_sin_banco"], delta="Verificar en banco", delta_color="inverse")
            m4.metric("⚠️ Diferencias", b["n_diferencias"])

            st.divider()
            st.markdown("### Cruce banco ↔ SAP ↔ CFDI")
            st.dataframe(
                b["tabla"],
                width="stretch",
                height=380,
                column_config={
                    "Cargo": st.column_config.NumberColumn(format="$%.2f"),
                    "CRUZE": st.column_config.NumberColumn(format="$%.2f"),
                    "DIF": st.column_config.NumberColumn(format="$%.2f"),
                },
            )
            st.info(
                "En el Módulo 2 completo, los estados de cuenta reales de cada banco "
                "(Banorte, Bancomer, Santander…) se cargan automáticamente cada mes y el "
                "cruce alimenta directo el bloque 1.1 del papel de trabajo.",
                icon="🧩",
            )

# ===========================================================================
# TAB 5 — INGRESOS SAP ↔ BANCA (MÓDULO 3 · RUTA)
# ===========================================================================
with tab5:
    st.subheader("📈 Integración de Ingresos: SAP ↔ Banca")
    st.markdown(
        f'<span style="background:{FRIOS_PURPLE_LIGHT};color:white;padding:3px 12px;'
        f'border-radius:12px;font-size:0.8rem;font-weight:700;">MÓDULO 3 · DISPONIBLE EN LA RUTA INTEGRAL</span>',
        unsafe_allow_html=True,
    )
    st.markdown("")
    st.markdown(
        "El tercer módulo cierra el ciclo completo aplicando el **mismo motor de "
        "conciliación** al lado de los ingresos:"
    )
    st.markdown(
        "- 📄 Lectura de **CFDIs emitidos** y de los **cobros registrados en SAP** (cuentas por cobrar)\n"
        "- 🏦 Cruce automático contra los **depósitos de los estados de cuenta**\n"
        "- 🔎 Detección de **depósitos no identificados**, cobros sin depósito y diferencias\n"
        "- 📋 **Papel de trabajo de ingresos**, equivalente al del Anexo 11, exportable a Excel\n"
        "- 💬 Consultas en lenguaje natural: *\"¿qué depósitos no están facturados?\"*"
    )
    st.info(
        "Este módulo se construye sobre la conexión a SAP y el motor de cruce de los "
        "Módulos 1 y 2 — por eso en la ruta integral toma solo 5-6 semanas. "
        "Para activarlo necesitamos los exports de CFDIs emitidos y cobros del periodo.",
        icon="🗺️",
    )
